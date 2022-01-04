import requests
from bs4 import BeautifulSoup
# to use webdriver you need chromedriver.exe in the same folder as the .py script
from selenium import webdriver
import time
from datetime import date
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from datetime import date
import os
# Selenium Implementation

def get_html_source(url):
    # time_wait = input("Enter time(in s) to wait for you to put in pincode: ")
    browser = webdriver.Edge(r"msedgedriver.exe")
    # get web page
    browser.get(url)
    # execute script to scroll down the page
    time.sleep(int(40))
    print("Starting_scrape")
    browser.get(url)
    show_more_button = browser.find_element(By.CSS_SELECTOR, 'button[ng-click="vm.pagginator.showmorepage()"]')
    # time.sleep(600)
    for i in range(0,200):
        browser.execute_script("window.scrollTo(0, 8/9*document.body.scrollHeight);var lenOfPage=document.body.scrollHeight;return lenOfPage;")
        try:
            time.sleep(3)
            show_more_button.click()
        except:
            break
    time.sleep(10)
    return browser.page_source


soup = BeautifulSoup(get_html_source(url="https://www.bigbasket.com/cl/fruits-vegetables/"),"lxml")

product_list = soup.find_all(attrs={"qa": "product"})
final_data = []
pincode = soup.find_all("span", attrs={"ng-bind":"vm.user.currentAddress.address_display_name"})[0].text

for product_div in product_list:
    product_name = product_div.find_next(attrs={"qa": "product_name"})
    product_name = product_name.find_next("a").text

    weight = product_div.find_next("span",attrs={"ng-bind": "vm.selectedProduct.w"}).text
    price = product_div.find_next("span",attrs={"class":"discnt-price"}).text
    # price = price[3:]
    try:
        price = float(price)
    except :
        price = float(price.split(" ")[-1])
        
    final_data.append({
        "name": product_name,
        "weight":weight,
        "price": price,
    })
print(len(final_data), "no. of elements stored")

# Excel Code
excel_filename= f"{pincode}_{date.today().strftime('%d-%m-%Y')}.xlsx"

try:
    wb = load_workbook(excel_filename)
except FileNotFoundError:
    wb = Workbook()
    wb.save(excel_filename)

ws = wb.active
ws.cell(row=1, column=3).value = date.today().strftime("%d/%m/%Y")
ws.cell(row=2, column=2).value = "Price"
ws.cell(row=2, column=3).value = "Weight/pc(s)"

for index_1, dict in enumerate(final_data):
    
    ws.cell(row=3+index_1, column=1).value = dict["name"]
    ws.cell(row=3+index_1, column=2).value = dict["price"]
    ws.cell(row=3+index_1, column=3).value = dict["weight"]
    
wb.save(excel_filename)
print("Excel Created")
os.system("PAUSE")