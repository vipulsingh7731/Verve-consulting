import requests
from bs4 import BeautifulSoup
# to use webdriver you need chromedriver.exe in the same folder as the .py script
from selenium import webdriver
import time
from datetime import date
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from datetime import date
import os

# Selenium Implementation

def get_html_source(url):
    # time_wait = input("Enter time(in s) to wait for you to put in pincode: ")
    browser = webdriver.Chrome()
    # get web page
    browser.get(url)
    # execute script to scroll down the page
    # time.sleep(int(40))
    print("Starting_scrape")    
    browser.execute_script("window.scrollTo(0, 8/9*document.body.scrollHeight);var lenOfPage=document.body.scrollHeight;return lenOfPage;")
    time.sleep(10)
    return browser.page_source


print("Default URL used if nothing is provided -- https://market.todaypricerates.com/Bhubaneswar-vegetables-price-in-Odisha")
cond = False
while cond == False:
    the_input = input("Enter URL from market.todaypricerates.com that you want to scrape(Press Enter to use default):")
    if the_input=="":
        the_input ="https://market.todaypricerates.com/Bhubaneswar-vegetables-price-in-Odisha"
    try:
        url_split2 = the_input.split("/")
        url_split2 = url_split2[2]
    except:
        pass
    if url_split2 =="market.todaypricerates.com":
        cond = True
    else:
        print("Please Enter a valid URL!!!")
del cond,url_split2


soup = BeautifulSoup(get_html_source(url=the_input), "lxml")

table = soup.select("body > div.single-product-area > div.container > div > div.col-md-8 > div > div.Table")[0]


heading = table.find_next("div")
data = heading.find_next_siblings("div")
all_data = []
for row in data:
    vegetable_name = row.find_next("div")
    siblings = vegetable_name.find_next_siblings("div")
    unit = siblings[0]
    market_price = float(siblings[1].text.split(" ")[-1])
    retail_price = ( float(siblings[2].text.split(" ")[2]) + float(siblings[2].text.split(" ")[4]) ) / 2
    shopping_mall = ( float(siblings[3].text.split(" ")[2]) + float(siblings[3].text.split(" ")[4]) ) / 2
    all_data.append({
        "vegetable_name":vegetable_name.text,
        "unit":unit.text,
        "market_price":market_price,
        "retail_price":retail_price,
        "shopping_mall":shopping_mall
    })

# excel code
excel_filename= f"{table.find_previous().text}.xlsx"

try:
    wb = load_workbook(excel_filename)
except FileNotFoundError:
    wb = Workbook()
    wb.save(excel_filename)

ws = wb.active
maxrow=ws.max_row+1

unit_col=ws.max_column+1
date_col= unit_col
market_col=unit_col+1
retail_col = market_col+1
shopping_col = retail_col+1

no_of_col_perdate = 4
if date_col>no_of_col_perdate:
    if ws.cell(row=1, column=date_col- no_of_col_perdate).value == date.today().strftime("%d/%m/%Y"):
        unit_col = unit_col - no_of_col_perdate
        market_col = market_col - no_of_col_perdate
        retail_col = retail_col - no_of_col_perdate
        shopping_col = shopping_col - no_of_col_perdate
        date_col = date_col - no_of_col_perdate
ws.cell(row=1, column=date_col).value = date.today().strftime("%d/%m/%Y")
ws.cell(row=2, column=unit_col).value = "Unit"
ws.cell(row=2, column=market_col).value = "Market Price"
ws.cell(row=2, column=retail_col).value = "Retail Price"
ws.cell(row=2, column=shopping_col).value = "Shopping Mall Price"
for index_1, dict in enumerate(all_data):
    c=1
    for i in range(maxrow-1):
        if ws.cell(row=i+1, column=1).value == dict["vegetable_name"]:
            c=0
            ws.cell(row=i+1, column=unit_col).value = dict["unit"]
            ws.cell(row=i+1, column=market_col).value = dict["market_price"]
            ws.cell(row=i+1, column=retail_col).value = dict["retail_price"]
            ws.cell(row=i+1, column=shopping_col).value = dict["shopping_mall"]
    if c==1:
        ws.cell(row=ws.max_row+1, column=1).value = dict["vegetable_name"]
        ws.cell(row=ws.max_row, column=unit_col).value = dict["unit"]
        ws.cell(row=ws.max_row, column=market_col).value = dict["market_price"]
        ws.cell(row=ws.max_row, column=retail_col).value = dict["retail_price"]
        ws.cell(row=ws.max_row, column=shopping_col).value = dict["shopping_mall"]
        
wb.save(excel_filename)
print("Excel Created!!!")
os.system("PAUSE")